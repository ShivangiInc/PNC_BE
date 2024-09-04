import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from bson import ObjectId
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.views import View
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
from .models import table_data, deleted_columns
from .services import (
    process_excel_file,
    insert_records,
    clear_existing_records,
    clear_deleted_columns,
    fetch_all_deleted_column_names,
    fetch_all_deleted_by_admin_column_names,
    fetch_all_records,
    process_csv_file,
    process_excel_file,
    process_tsv_file,
    fetch_all_deleted_by_admin_record_names,
    fetch_all_rejected_by_admin_column_names,
    fetch_all_rejected_by_admin_record_names,


)

def plot_chart():
    file_path = r"C:\Users\bhawna.atrish\Downloads\chartData.xlsx"
    df = pd.read_excel(file_path, "Sheet1")
    month = df["Month"]
    profit = df["Profit"]
    plt.figure(figsize=(10, 5))
    plt.bar(month, profit, color="Skyblue")
    plt.show()

class ExcelUploadView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Handle POST requests to upload an Excel, CSV, or TSV file and replace existing data in MongoDB.
        http://localhost:8000/api/upload/
        """
        if "file" not in request.FILES:
            return Response(
                {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
            )

        uploaded_file = request.FILES["file"]
        file_name = uploaded_file.name

        try:
            if file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                # Process Excel file
                records = process_excel_file(uploaded_file)
            elif file_name.endswith(".csv"):
                # Process CSV file
                records = process_csv_file(uploaded_file)
            elif file_name.endswith(".tsv"):
                # Process TSV file
                records = process_tsv_file(uploaded_file)
            else:
                return Response(
                    {
                        "error": "Unsupported file format. Please upload an Excel, CSV, or TSV file."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Clear existing data and insert new records
            clear_deleted_columns()
            clear_existing_records()
            insert_records(records)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"message": "Data successfully replaced in MongoDB"},
            status=status.HTTP_201_CREATED,
        )

class ExcelDataView(APIView):
    """
    Handle GET requests to retrieve data from MongoDB and return it as a list of dictionaries.
    http://localhost:8000/api/data/
    """

    def get(self, request, *args, **kwargs):
        try:
            # Fetch all records
            records = fetch_all_records()

            # Fetch deleted columns
            deleted_columns = fetch_all_deleted_column_names()
            deleted_by_admin_columns = fetch_all_deleted_by_admin_column_names()
            deleted_by_admin_records = fetch_all_deleted_by_admin_record_names()
            rejected_by_admin_columns = fetch_all_rejected_by_admin_column_names()
            rejected_by_admin_records = fetch_all_rejected_by_admin_record_names()

            # Combine the data into a single response
            response_data = {
                "records": records,
                "deleted_columns": deleted_columns,
                "deleted_by_admin_columns": deleted_by_admin_columns,
                "deleted_by_admin_records": deleted_by_admin_records,
                "rejected_by_admin_columns": rejected_by_admin_columns,
                "rejected_by_admin_records": rejected_by_admin_records,

            }

        except ValueError as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(response_data, status=status.HTTP_200_OK)


class ModifyRecordView(APIView):
    def post(self, request, record_id=None, *args, **kwargs):
        """
        Update a specific row based on record_id, or create a new row if record_id is not found.
        Create new record: http://localhost:8000/api/create_or_update_record/
        Update record: http://localhost:8000/api/create_or_update_record/66b9fb790b2700bfd39597b8/
        """
        try:
            update_data = request.data

            if record_id:
                # Update an existing record
                object_id = ObjectId(record_id)
                current_document = table_data.find_one({"_id": object_id})

                if not current_document:
                    return Response(
                        {"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND
                    )

                # Filter out any fields not in the current document
                allowed_fields = set(current_document.keys())
                filtered_update_data = {
                    key: value
                    for key, value in update_data.items()
                    if key in allowed_fields
                }

                if not filtered_update_data:
                    return Response(
                        {"error": "No valid fields to update"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                result = table_data.update_one(
                    {"_id": object_id}, {"$set": filtered_update_data}
                )

                if result.matched_count == 0:
                    return Response(
                        {"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND
                    )

                return Response(
                    {"message": "Record updated successfully"},
                    status=status.HTTP_200_OK,
                )
            else:
                # Create a new record
                result = table_data.insert_one(update_data)
                return Response(
                    {
                        "message": "New row created successfully",
                        "id": str(result.inserted_id),
                    },
                    status=status.HTTP_201_CREATED,
                )

        except Exception as e:
            return Response(
                {"error": f"Error processing request in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def delete(self, request, record_id, *args, **kwargs):
        """
        Soft-delete a specific row by marking it as deleted.
        http://localhost:8000/api/create_or_update_record/66b9fb790b2700bfd39597ba/
        """
        try:
            result = table_data.update_one(
                {"_id": ObjectId(record_id)}, {"$set": {"is_deleted": True}}
            )

            if result.matched_count == 0:
                return Response(
                    {"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND
                )

            return Response(
                {"message": "Record marked as deleted successfully"},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error marking record as deleted in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AddColumnView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Add a new column to every document in the MongoDB collection.
        http://localhost:8000/api/add-column/
        """
        try:
            column_name = request.data.get("column_name")
            if not column_name:
                return Response(
                    {"error": "Column name is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            result = table_data.update_many({}, {"$set": {column_name: None}})

            return Response(
                {
                    "message": f"Column '{column_name}' added to {result.modified_count} documents"
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error adding column to MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class SoftDeleteColumnView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Mark a column as soft-deleted by adding an entry to a separate collection.
        The original column in the main collection will remain unaffected.
        Endpoint: http://localhost:8000/api/soft-delete-column/
        """
        try:
            column_name = request.data.get("column_name")
            if not column_name:
                return Response(
                    {"error": "Column name is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Check if the column is already in the `deleted_columns` collection
            existing_entry = deleted_columns.find_one({"column_name": column_name})

            if existing_entry:
                # If it exists, update the "is_deleted" field to True
                deleted_columns.update_one(
                    {"column_name": column_name}, {"$set": {"is_deleted": True}}
                )
            else:
                # If it does not exist, create a new document for the column
                deleted_columns.insert_one(
                    {"column_name": column_name, "is_deleted": True}
                )

            return Response(
                {
                    "message": f"Column '{column_name}' has been marked as deleted in the tracking collection."
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error marking column as deleted in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class RenameColumnView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Rename a column in every document in the collection.
        Takes 'old_column_name' and 'new_column_name' as input.
        http://localhost:8000/api/rename-column/
        """
        try:
            old_column_name = request.data.get("old_column_name")
            new_column_name = request.data.get("new_column_name")

            if not old_column_name or not new_column_name:
                return Response(
                    {
                        "error": "Both 'old_column_name' and 'new_column_name' are required"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            result = table_data.update_many(
                {},
                [
                    {"$set": {new_column_name: "$" + old_column_name}},
                    {"$unset": old_column_name},
                ],
            )

            return Response(
                {
                    "message": f"Column '{old_column_name}' renamed to '{new_column_name}' in {result.modified_count} documents"
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error renaming column in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

def generate_chart(df):
    # 9. Create the bar chart using Matplotlib
    plt.figure(figsize=(10, 5))
    plt.bar(
        df["EODBalance-14Aug"], df["Account ID"], color="Skyblue"
    )  # Adjust the columns as per your data
    plt.title("Monthly Profit")
    plt.xlabel("Month")
    plt.ylabel("Profit")

    # 10. Save the plot to a BytesIO object
    img_data = BytesIO()
    plt.savefig(img_data, format="png")
    img_data.seek(0)  # Rewind the data to the beginning

    # 11. Close the plot to free up memory
    plt.close()

    return img_data


class ExcelExportView(View):
    def get(self, request, *args, **kwargs):
        # Fetch data from the MongoDB collection
        mongo_data = list(table_data.find({}, {'_id': 0}))  # Query without '_id' field

        # Convert the data to a DataFrame
        df = pd.DataFrame(mongo_data)

        # Save the DataFrame to a BytesIO object
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        buffer.seek(0)
        # plot_chart()

        # Create the HTTP response with the appropriate content type and headers
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

        return response

    # def get(self, request, *args, **kwargs):
    #     # 1. Fetch data from the MongoDB collection
    #     mongo_data = list(table_data.find({}, {"_id": 0}))  # Query without '_id' field

    #     # 2. Convert the data to a DataFrame
    #     df = pd.DataFrame(mongo_data)

    #     # 3. Save the DataFrame to a BytesIO object (Excel file in memory)
    #     buffer = BytesIO()
    #     with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    #         df.to_excel(writer, index=False, sheet_name="Sheet1")
    #     buffer.seek(0)

    #     # 4. Load the workbook to embed the chart
    #     workbook = load_workbook(buffer)
    #     worksheet = workbook.active

    #     # 5. Generate the chart using Matplotlib
    #     chart_img_data = generate_chart(df)

    #     # 6. Embed the chart image into the Excel file
    #     img = Image(chart_img_data)
    #     img.anchor = "E5"  # Position the chart at cell E5 (adjust as needed)
    #     worksheet.add_image(img)

    #     # 7. Save the workbook to the buffer
    #     buffer = BytesIO()
    #     workbook.save(buffer)
    #     buffer.seek(0)

    #     # 8. Create the HTTP response with the appropriate content type and headers
    #     response = HttpResponse(
    #         buffer.getvalue(),
    #         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     )
    #     response["Content-Disposition"] = 'attachment; filename="data_with_chart.xlsx"'

    #     return response


# def get(self, request, *args, **kwargs):
#     # Fetch data from the MongoDB collection
#     mongo_data = list(table_data.find({}, {'_id': 0}))  # Query without '_id' field

#     # Convert the data to a DataFrame
#     df = pd.DataFrame(mongo_data)

#     # Save the DataFrame to a BytesIO object
#     buffer = BytesIO()
#     with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Sheet1')

#         # Access the openpyxl workbook and worksheet
#         workbook = writer.book
#         worksheet = writer.sheets['Sheet1']

#         # Create a clustered column chart
#         chart = BarChart()
#         chart.type = "col"
#         chart.style = 10
#         chart.grouping = "clustered"
#         chart.title = "Clustered Column Chart Example"
#         chart.y_axis.title = 'Values'
#         chart.x_axis.title = 'Categories'

#         # Set x-axis categories (from the first column, usually categorical data)
#         categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(df) + 1)

#         # Add each numeric column as a separate series in the chart
#         for idx in range(2, len(df.columns) + 1):
#             values = Reference(worksheet, min_col=idx, min_row=2, max_row=len(df) + 1)
#             series = Series(values, title_from_data=True)
#             chart.append(series)

#         # Set categories (X-axis)
#         chart.set_categories(categories)

#         # Position the chart on the worksheet
#         worksheet.add_chart(chart, "E5")

#     buffer.seek(0)

#     # Create the HTTP response with the appropriate content type and headers
#     response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="data_with_clustered_chart.xlsx"'

#     return response
# def get(self, request, *args, **kwargs):
#     # Example data; replace with your MongoDB query result
#     data = list(table_data.find({}, {'_id': 0}))  # Query without '_id' field

#     df = pd.DataFrame(data)

#     # Create an Excel workbook and add a worksheet
#     workbook = Workbook()
#     worksheet = workbook.active
#     worksheet.title = "Sheet1"

#     # Write DataFrame data to Excel
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
#         for c_idx, value in enumerate(row, 1):
#             worksheet.cell(row=r_idx, column=c_idx, value=value)

#     # Create a clustered column chart
#     chart = BarChart()
#     chart.type = "col"
#     chart.grouping = "clustered"
#     chart.title = "Custom Clustered Column Chart"
#     chart.y_axis.title = 'Values'
#     chart.x_axis.title = 'MDMID'

#     # Manually define series for each column with specific formulas
#     series_formulas = [
#         f'=SERIES("Sum of EODBalance-14Aug",{{"MDMID18","MDMID14"}},{{149524.939708448,4738966.82449053}},1)',
#         f'=SERIES("Sum of EODBalance-15Aug",{{"MDMID18","MDMID14"}},{{145729.058052051,4737172.30504085}},2)',
#         f'=SERIES("Sum of EODBalance-17Aug",{{"MDMID18","MDMID14"}},{{143209.597556718,4741859.08356868}},3)',
#         f'=SERIES("Sum of EODBalance-18Aug",{{"MDMID18","MDMID14"}},{{142582.036089139,4738420.27698655}},4)',
#         f'=SERIES("Sum of EODBalance-16Aug",{{"MDMID18","MDMID14"}},{{145729.058052051,4737172.30504085}},5)',
#         f'=SERIES("Sum of 5-Day average",{{"MDMID18","MDMID14"}},{{145354.937891681,4738718.15902549}},6)',
#         f'=SERIES("Sum of Projected Balance",{{"MDMID18","MDMID14"}},{{116888,5727800}},7)',
#     ]

#     for formula in series_formulas:
#         series = Series(values=Reference(worksheet, min_col=2, min_row=2, max_row=len(df) + 1))
#         series.formula = formula
#         chart.series.append(series)

#     # Add the chart to the worksheet
#     worksheet.add_chart(chart, "E5")

#     # Save the workbook to a BytesIO object
#     buffer = BytesIO()
#     workbook.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response with the appropriate content type and headers
#     response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="custom_data_with_chart.xlsx"'

#     return response


class PdfExportView(View):
    def get(self, request, *args, **kwargs):
        # Fetch records from MongoDB
        records = fetch_all_records()

        # Convert records to DataFrame
        df = pd.DataFrame(records)

        # Drop the '_id' column if it exists
        if "_id" in df.columns:
            df = df.drop(columns=["_id"])

        # Convert DataFrame to a list of lists (rows)
        table_data = [df.columns.tolist()] + df.values.tolist()

        # Estimate the width of each column
        col_widths = [
            max([len(str(item)) for item in col]) * 0.1 * inch
            for col in zip(*table_data)
        ]

        # Calculate the total width of the table
        total_width = sum(col_widths)

        # Set page size dynamically based on the table width
        page_width = total_width + 2 * inch  # Add some padding
        page_size = (
            page_width,
            11 * inch,
        )  # Keep height standard (11 inches for A4 height)

        # Create a BytesIO buffer for the PDF
        pdf_buffer = BytesIO()

        # Create the PDF document and canvas
        doc = SimpleDocTemplate(pdf_buffer, pagesize=page_size)
        elements = []

        # Create a Table object from the data with calculated column widths
        table = Table(table_data, colWidths=col_widths)

        # Apply some table styling
        style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
        table.setStyle(style)

        # Add the table to the document elements
        elements.append(table)

        # Build the PDF document
        doc.build(elements)

        # Get PDF data from the buffer
        pdf = pdf_buffer.getvalue()
        pdf_buffer.close()

        # Create an HTTP response with the PDF content
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="table_data.pdf"'

        return response

class ColDeletionApprovedView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Soft delete columns approved by admin based on column names.
        """
        column_names = request.data.get("column_names", [])

        if not column_names:
            return Response(
                {"error": "No column names provided"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            update_results = []
            for column_name in column_names:
                # Update operation
                filter_query = {"column_name": column_name}
                update_operation = {"$set": {"deleted_by_admin": True}}
                result = deleted_columns.update_one(filter_query, update_operation)

                update_results.append({
                    "column_name": column_name,
                    "matched_count": result.matched_count,
                    "modified_count": result.modified_count
                })

            matched_count = sum(res["matched_count"] for res in update_results)
            modified_count = sum(res["modified_count"] for res in update_results)

            if matched_count == 0:
                return Response(
                    {"error": "No matching columns found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            return Response(
                {"message": f"{matched_count} column(s) marked as deleted successfully"},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error marking columns as deleted in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class ColDeletionRejectedView(APIView):
    def post(self, request, *args, **kwargs):
        """
        Remove 'is_deleted' field for multiple columns based on column names.
        """
        column_names = request.data.get("column_names", [])

        if not column_names:
            return Response(
                {"error": "No column names provided"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            update_results = []
            for column_name in column_names:
                # Update operation
                filter_query = {"column_name": column_name}
                update_operation = {"$unset": {"is_deleted": ""},"$set": {"deleted_by_admin": False}}
                result = deleted_columns.update_one(filter_query, update_operation)

                update_results.append({
                    "column_name": column_name,
                    "matched_count": result.matched_count,
                    "modified_count": result.modified_count
                })

            matched_count = sum(res["matched_count"] for res in update_results)
            modified_count = sum(res["modified_count"] for res in update_results)

            if matched_count == 0:
                return Response(
                    {"error": "No matching columns found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            return Response(
                {"message": f"{matched_count} column(s) updated successfully"},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error updating columns in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class RecordDeletionApproved(APIView):   
    def post(self, request, *args, **kwargs):
        """
        Mark multiple rows as deleted by admin.
        """
        record_ids = request.data.get("record_ids", [])
        if not record_ids:
            return Response(
                {"error": "No record IDs provided"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Convert string IDs to ObjectId
            object_ids = [ObjectId(record_id) for record_id in record_ids]
            filter_query = {"_id": {"$in": object_ids}}
            update_operation = {"$set": {"deleted_by_admin": True}}
            result = table_data.update_many(
               filter_query, update_operation
            )
            print(f"Matched {result.matched_count} documents and modified {result.modified_count} documents.")

            if result.matched_count == 0:
                return Response(
                    {"error": "No matching records found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            return Response(
                {"message": f"{result.matched_count} record(s) marked as deleted by admin successfully"},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error marking records as deleted by admin in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class RecordDeletionDisapproved(APIView):
    def post(self, request, *args, **kwargs):
        """
        Remove the 'is_deleted' field from multiple rows.
        """
        record_ids = request.data.get("record_ids", [])

        if not record_ids:
            return Response(
                {"error": "No record IDs provided"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Convert string IDs to ObjectId
            object_ids = [ObjectId(record_id) for record_id in record_ids]

            # Update operation to unset 'is_deleted' field
            filter_query = {"_id": {"$in": object_ids}}
            update_operation = {"$unset": {"is_deleted": ""},"$set": {"deleted_by_admin": False}}
            result = table_data.update_many(
                filter_query, update_operation
            )

            if result.matched_count == 0:
                return Response(
                    {"error": "No matching records found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            return Response(
                {"message": f"{result.matched_count} record(s) updated successfully"},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": f"Error updating records in MongoDB: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        
# class DeletionRecord(APIView):
#     def get(self, request, *args, **kwargs):

